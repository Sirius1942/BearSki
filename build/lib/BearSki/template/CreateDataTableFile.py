from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
from openpyxl.styles import  PatternFill

def createDTF(path):

  wb = Workbook()

  dest_filename = os.path.join(path,'testdata.xlsx')

  ws1 = wb.active
  ws1.title = "example"
  ws1['A1']='DataID'
  ws1['B1']='Remark'
  ws1['C1']='username'
  ws1['D1']='password'

  ws1['A2']='admin'
  ws1['B2']='系统管理员帐号'
  ws1['C2']='admin'
  ws1['D2']='agave@123456'

  ws1['A3']='e1'
  ws1['B3']='时间与随机数函数'
  ws1['C3']='${time.now}'
  ws1['D3']='${random.int(8)}'

  ws1['A4']='e2'
  ws1['B4']='cell文本转数据和json格式数据'
  ws1['C4']='$list{3,4.int,5}'
  ws1['D4']='$json{{"name":"123"}}'

  ws1.column_dimensions['A'].width = 10
  ws1.column_dimensions['B'].width = 30
  ws1.column_dimensions['C'].width = 25
  ws1.column_dimensions['D'].width = 25

  fill = PatternFill("solid", fgColor="1874CD")
  ws1['A1'].fill = fill
  ws1['B1'].fill = fill
  ws1['C1'].fill = fill
  ws1['D1'].fill = fill

  wb.save(filename = dest_filename)

if __name__ == '__main__':
	createDTF('')
